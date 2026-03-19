"""
Admin panel handlers - Stats, Users, Transactions, Export

This module handles:
- Detailed statistics display
- Users list export to Excel
- Payments list export to Excel
- Active/Removed users list
"""
from datetime import datetime
from io import BytesIO
from aiogram import F
from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, BufferedInputFile
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.database import AsyncSessionLocal
from client_bot.config import logger, TIMEZONE
from client_bot.utils.database import get_bot_info, get_client_language
from locales import get_text as get_admin_text


def register_admin_handlers(dp, bot, owner_id: int, bot_token: str, bot_db_id: int):
    """Register admin panel handlers"""

    @dp.callback_query(F.data == "admin_stats")
    async def show_admin_stats(callback: CallbackQuery):
        """Show detailed statistics"""
        if callback.from_user.id != owner_id:
            await callback.answer(get_admin_text("no_permission", "uz"), show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            try:
                # Get client language
                lang = await get_client_language(session, owner_id)

                # Get bot info
                bot_info = await get_bot_info(session, bot_token)

                # Today's date for filtering
                today = datetime.now(TIMEZONE).date()
                today_start = datetime.combine(today, datetime.min.time())

                # Today's new users (per-bot)
                today_query = text("""
                    SELECT COUNT(*) FROM users
                    WHERE admin_id = :admin_id AND bot_id = :bot_id AND created_at >= :today_start
                """)
                today_result = await session.execute(today_query, {"admin_id": owner_id, "bot_id": bot_db_id, "today_start": today_start})
                today_users = today_result.scalar() or 0

                # Total users (per-bot)
                total_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id AND bot_id = :bot_id")
                total_result = await session.execute(total_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                total_users = total_result.scalar() or 0

                # Active users (per-bot)
                active_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id AND bot_id = :bot_id AND status = 'active'")
                active_result = await session.execute(active_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                active_users = active_result.scalar() or 0

                # Removed users (per-bot)
                removed_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id AND bot_id = :bot_id AND status = 'removed'")
                removed_result = await session.execute(removed_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                removed_users = removed_result.scalar() or 0

                # Total revenue (approved)
                revenue_query = text("""
                    SELECT COALESCE(SUM(amount), 0) FROM transactions
                    WHERE admin_id = :admin_id AND role = 'plan_purchase' AND status = 'approved'
                """)
                revenue_result = await session.execute(revenue_query, {"admin_id": owner_id})
                total_revenue = float(revenue_result.scalar() or 0)

                # Approved amount
                approved_query = text("""
                    SELECT COALESCE(SUM(amount), 0) FROM transactions
                    WHERE admin_id = :admin_id AND status = 'approved'
                """)
                approved_result = await session.execute(approved_query, {"admin_id": owner_id})
                approved_amount = float(approved_result.scalar() or 0)

                # Rejected amount
                rejected_query = text("""
                    SELECT COALESCE(SUM(amount), 0) FROM transactions
                    WHERE admin_id = :admin_id AND status = 'rejected'
                """)
                rejected_result = await session.execute(rejected_query, {"admin_id": owner_id})
                rejected_amount = float(rejected_result.scalar() or 0)

                # Pending count
                pending_query = text("""
                    SELECT COUNT(*) FROM transactions
                    WHERE admin_id = :admin_id AND status = 'pending'
                """)
                pending_result = await session.execute(pending_query, {"admin_id": owner_id})
                pending_count = pending_result.scalar() or 0

                # Users by duration (subscription type) (per-bot)
                monthly_query = text("""
                    SELECT COUNT(*) FROM users
                    WHERE admin_id = :admin_id AND bot_id = :bot_id AND duration = '1 oy' AND status = 'active'
                """)
                monthly_result = await session.execute(monthly_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                monthly_users = monthly_result.scalar() or 0

                yearly_query = text("""
                    SELECT COUNT(*) FROM users
                    WHERE admin_id = :admin_id AND bot_id = :bot_id AND duration = '1 yil' AND status = 'active'
                """)
                yearly_result = await session.execute(yearly_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                yearly_users = yearly_result.scalar() or 0

                unlimited_query = text("""
                    SELECT COUNT(*) FROM users
                    WHERE admin_id = :admin_id AND bot_id = :bot_id AND duration = 'cheksiz' AND status = 'active'
                """)
                unlimited_result = await session.execute(unlimited_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                unlimited_users = unlimited_result.scalar() or 0

                # Build stats message with translations
                currency = get_admin_text("currency", lang)
                people = get_admin_text("people", lang)
                pcs = get_admin_text("pcs", lang)

                stats_msg = f"📊 {get_admin_text('stats_title', lang)}\n\n"

                if bot_info:
                    stats_msg += f"🤖 {get_admin_text('bot_label', lang)} @{bot_info.get('bot_username', 'N/A')}\n"
                    

                stats_msg += f"👥 {get_admin_text('users_section', lang)}\n"
                stats_msg += f"  📅 {get_admin_text('today_joined', lang)}: {today_users}\n"
                stats_msg += f"  📊 {get_admin_text('total', lang)}: {total_users}\n"
                stats_msg += f"  ✅ {get_admin_text('active', lang)}: {active_users}\n"
                stats_msg += f"  ❌ {get_admin_text('removed', lang)}: {removed_users}\n\n"

                stats_msg += f"💰 {get_admin_text('financial_section', lang)}\n"
                stats_msg += f"  💵 {get_admin_text('total_revenue', lang)}: {total_revenue:,.0f} {currency}\n"
                stats_msg += f"  ✅ {get_admin_text('approved_amount', lang)}: {approved_amount:,.0f} {currency}\n"
                stats_msg += f"  ❌ {get_admin_text('rejected_amount', lang)}: {rejected_amount:,.0f} {currency}\n"
                stats_msg += f"  ⏳ {get_admin_text('pending_count', lang)}: {pending_count} {pcs}\n\n"

                stats_msg += f"📋 {get_admin_text('subscription_section', lang)}\n"
                stats_msg += f"  📆 {get_admin_text('monthly_sub', lang)}: {monthly_users} {people}\n"
                stats_msg += f"  📅 {get_admin_text('yearly_sub', lang)}: {yearly_users} {people}\n"
                stats_msg += f"  ♾ {get_admin_text('unlimited_sub', lang)}: {unlimited_users} {people}\n"

                back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text=f"🔙 {get_admin_text('btn_back', lang)}", callback_data="admin_back")]
                ])

                await callback.message.edit_text(stats_msg, parse_mode="HTML", reply_markup=back_keyboard)
                await callback.answer()

            except Exception as e:
                logger.error(f"Stats error: {e}")
                await callback.answer(get_admin_text("error_occurred", "uz"), show_alert=True)

    @dp.callback_query(F.data == "admin_users_excel")
    async def export_users_excel(callback: CallbackQuery):
        """Export users list to Excel"""
        if callback.from_user.id != owner_id:
            await callback.answer(get_admin_text("no_permission", "uz"), show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            # Get client language
            lang = await get_client_language(session, owner_id)

            await callback.answer(get_admin_text("users_preparing", lang))

            try:
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Users"

                # Headers with translations
                headers = [
                    get_admin_text("excel_num", lang),
                    get_admin_text("excel_user_id", lang),
                    get_admin_text("excel_username", lang),
                    get_admin_text("excel_name", lang),
                    get_admin_text("excel_duration", lang),
                    get_admin_text("excel_status", lang),
                    get_admin_text("excel_balance", lang),
                    get_admin_text("excel_date", lang)
                ]
                ws.append(headers)

                # Style headers
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Get users data (per-bot)
                users_query = text("""
                    SELECT id, user_id, username, name, duration, status, balance, created_at
                    FROM users
                    WHERE admin_id = :admin_id AND bot_id = :bot_id
                    ORDER BY created_at DESC
                """)
                users_result = await session.execute(users_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                users = users_result.fetchall()

                for idx, user in enumerate(users, 1):
                    user_dict = dict(user._mapping)

                    # Translate status
                    status_raw = user_dict.get('status') or 'N/A'
                    if status_raw == 'active':
                        status_text = get_admin_text("active", lang)
                    elif status_raw == 'removed':
                        status_text = get_admin_text("removed", lang)
                    else:
                        status_text = status_raw

                    ws.append([
                        idx,
                        user_dict.get('user_id'),
                        user_dict.get('username') or 'N/A',
                        user_dict.get('name') or 'N/A',
                        user_dict.get('duration') or 'N/A',
                        status_text,
                        float(user_dict.get('balance') or 0),
                        user_dict.get('created_at').strftime('%Y-%m-%d %H:%M') if user_dict.get('created_at') else 'N/A'
                    ])

                # Adjust column widths
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 18

                # Save to BytesIO
                excel_file = BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)

                # Send file
                file_name = f"users_{datetime.now(TIMEZONE).strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_input = BufferedInputFile(excel_file.read(), filename=file_name)

                pcs = get_admin_text("pcs", lang)
                await callback.message.answer_document(
                    document=excel_input,
                    caption=f"📋 {get_admin_text('users_list_title', lang)}\n\n👥 {get_admin_text('total_users', lang)}: {len(users)} {pcs}",
                    parse_mode="HTML"
                )

            except Exception as e:
                logger.error(f"Users Excel export error: {e}")
                await callback.message.answer(f"❌ {get_admin_text('excel_error', lang)}")

    @dp.callback_query(F.data == "admin_payments_excel")
    async def export_payments_excel(callback: CallbackQuery):
        """Export payments list to Excel"""
        if callback.from_user.id != owner_id:
            await callback.answer(get_admin_text("no_permission", "uz"), show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            # Get client language
            lang = await get_client_language(session, owner_id)

            await callback.answer(get_admin_text("payments_preparing", lang))

            try:
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Payments"

                # Headers with translations
                headers = [
                    get_admin_text("excel_num", lang),
                    get_admin_text("excel_trans_id", lang),
                    get_admin_text("excel_user_id", lang),
                    get_admin_text("excel_username", lang),
                    get_admin_text("excel_amount", lang),
                    get_admin_text("excel_type", lang),
                    get_admin_text("excel_status", lang),
                    get_admin_text("excel_trans_date", lang)
                ]
                ws.append(headers)

                # Style headers
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Get transactions data
                trans_query = text("""
                    SELECT id, user_id, username, amount, role, status, created_at
                    FROM transactions
                    WHERE admin_id = :admin_id
                    ORDER BY created_at DESC
                """)
                trans_result = await session.execute(trans_query, {"admin_id": owner_id})
                transactions = trans_result.fetchall()

                for idx, trans in enumerate(transactions, 1):
                    trans_dict = dict(trans._mapping)

                    # Translate role
                    role_raw = trans_dict.get('role')
                    if role_raw == 'plan_purchase':
                        role_text = get_admin_text("trans_plan_purchase", lang)
                    else:
                        role_text = get_admin_text("trans_topup", lang)

                    # Translate status
                    status_raw = trans_dict.get('status')
                    if status_raw == 'approved':
                        status_text = get_admin_text("status_approved", lang)
                    elif status_raw == 'rejected':
                        status_text = get_admin_text("status_rejected", lang)
                    elif status_raw == 'pending':
                        status_text = get_admin_text("status_pending", lang)
                    else:
                        status_text = status_raw or 'N/A'

                    ws.append([
                        idx,
                        trans_dict.get('id'),
                        trans_dict.get('user_id'),
                        trans_dict.get('username') or 'N/A',
                        float(trans_dict.get('amount') or 0),
                        role_text,
                        status_text,
                        trans_dict.get('created_at').strftime('%Y-%m-%d %H:%M') if trans_dict.get('created_at') else 'N/A'
                    ])

                # Adjust column widths
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 18
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 18

                # Save to BytesIO
                excel_file = BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)

                # Send file
                file_name = f"payments_{datetime.now(TIMEZONE).strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_input = BufferedInputFile(excel_file.read(), filename=file_name)

                pcs = get_admin_text("pcs", lang)
                await callback.message.answer_document(
                    document=excel_input,
                    caption=f"💳 {get_admin_text('payments_list_title', lang)}\n\n📊 {get_admin_text('total_payments', lang)}: {len(transactions)} {pcs}",
                    parse_mode="HTML"
                )

            except Exception as e:
                logger.error(f"Payments Excel export error: {e}")
                await callback.message.answer(f"❌ {get_admin_text('excel_error', lang)}")

    @dp.callback_query(F.data == "admin_active_users")
    async def show_active_users(callback: CallbackQuery):
        """Show last 10 active users"""
        if callback.from_user.id != owner_id:
            await callback.answer(get_admin_text("no_permission", "uz"), show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            try:
                # Get client language
                lang = await get_client_language(session, owner_id)

                # Get active users (last 10, per-bot)
                users_query = text("""
                    SELECT user_id, username, name, duration, created_at
                    FROM users
                    WHERE admin_id = :admin_id AND bot_id = :bot_id AND status = 'active'
                    ORDER BY created_at DESC
                    LIMIT 10
                """)
                result = await session.execute(users_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                users = result.fetchall()

                # Get total count (per-bot)
                count_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id AND bot_id = :bot_id AND status = 'active'")
                count_result = await session.execute(count_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                total_count = count_result.scalar() or 0

                pcs = get_admin_text("pcs", lang)
                users_msg = f"✅ {get_admin_text('active_users_title', lang)} ({total_count} {pcs})\n\n"

                if not users:
                    users_msg += get_admin_text("no_active_users", lang)
                else:
                    for idx, user in enumerate(users, 1):
                        user_dict = dict(user._mapping)
                        users_msg += f"{idx}. <b>{user_dict.get('name', 'N/A')}</b>\n"
                        users_msg += f"   👤 @{user_dict.get('username') or 'N/A'}\n"
                        users_msg += f"   📋 {get_admin_text('duration_label', lang)}: {user_dict.get('duration', 'N/A')}\n"
                        if user_dict.get('created_at'):
                            users_msg += f"   📅 {user_dict['created_at'].strftime('%d.%m.%Y')}\n"
                        users_msg += "\n"

                    if total_count > 10:
                        users_msg += f"\n{get_admin_text('last_10_shown', lang, total=total_count)}"

                back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text=f"🔙 {get_admin_text('btn_back', lang)}", callback_data="admin_back")]
                ])

                await callback.message.edit_text(users_msg, parse_mode="HTML", reply_markup=back_keyboard)
                await callback.answer()

            except Exception as e:
                logger.error(f"Active users error: {e}")
                await callback.answer(get_admin_text("error_occurred", "uz"), show_alert=True)

    @dp.callback_query(F.data == "admin_removed_users")
    async def show_removed_users(callback: CallbackQuery):
        """Show last 10 removed users"""
        if callback.from_user.id != owner_id:
            await callback.answer(get_admin_text("no_permission", "uz"), show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            try:
                # Get client language
                lang = await get_client_language(session, owner_id)

                # Get removed users (last 10, per-bot)
                users_query = text("""
                    SELECT user_id, username, name, duration, created_at
                    FROM users
                    WHERE admin_id = :admin_id AND bot_id = :bot_id AND status = 'removed'
                    ORDER BY created_at DESC
                    LIMIT 10
                """)
                result = await session.execute(users_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                users = result.fetchall()

                # Get total count (per-bot)
                count_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id AND bot_id = :bot_id AND status = 'removed'")
                count_result = await session.execute(count_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                total_count = count_result.scalar() or 0

                pcs = get_admin_text("pcs", lang)
                users_msg = f"❌ {get_admin_text('removed_users_title', lang)} ({total_count} {pcs})\n\n"

                if not users:
                    users_msg += get_admin_text("no_removed_users", lang)
                else:
                    for idx, user in enumerate(users, 1):
                        user_dict = dict(user._mapping)
                        users_msg += f"{idx}. <b>{user_dict.get('name', 'N/A')}</b>\n"
                        users_msg += f"   👤 @{user_dict.get('username') or 'N/A'}\n"
                        users_msg += f"   📋 {get_admin_text('duration_label', lang)}: {user_dict.get('duration', 'N/A')}\n"
                        if user_dict.get('created_at'):
                            users_msg += f"   📅 {user_dict['created_at'].strftime('%d.%m.%Y')}\n"
                        users_msg += "\n"

                    if total_count > 10:
                        users_msg += f"\n{get_admin_text('last_10_shown', lang, total=total_count)}"

                back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text=f"🔙 {get_admin_text('btn_back', lang)}", callback_data="admin_back")]
                ])

                await callback.message.edit_text(users_msg, parse_mode="HTML", reply_markup=back_keyboard)
                await callback.answer()

            except Exception as e:
                logger.error(f"Removed users error: {e}")
                await callback.answer(get_admin_text("error_occurred", "uz"), show_alert=True)

    @dp.callback_query(F.data == "admin_back")
    async def admin_back(callback: CallbackQuery):
        """Return to admin panel"""
        if callback.from_user.id != owner_id:
            await callback.answer(get_admin_text("no_permission", "uz"), show_alert=True)
            return

        async with AsyncSessionLocal() as session:
            try:
                # Get client language
                lang = await get_client_language(session, owner_id)

                # Get bot info
                bot_info = await get_bot_info(session, bot_token)

                # Get statistics (per-bot)
                total_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id AND bot_id = :bot_id")
                total_result = await session.execute(total_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                total_users = total_result.scalar() or 0

                active_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id AND bot_id = :bot_id AND status = 'active'")
                active_result = await session.execute(active_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                active_users = active_result.scalar() or 0

                removed_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id AND bot_id = :bot_id AND status = 'removed'")
                removed_result = await session.execute(removed_query, {"admin_id": owner_id, "bot_id": bot_db_id})
                removed_users = removed_result.scalar() or 0

                revenue_query = text("""
                    SELECT COALESCE(SUM(amount), 0) FROM transactions
                    WHERE admin_id = :admin_id AND role = 'plan_purchase' AND status = 'approved'
                """)
                revenue_result = await session.execute(revenue_query, {"admin_id": owner_id})
                total_revenue = float(revenue_result.scalar() or 0)

            except Exception as e:
                logger.error(f"Stats error: {e}")
                total_users = 0
                active_users = 0
                removed_users = 0
                total_revenue = 0
                bot_info = None
                lang = "uz"

        # Build admin panel with translations
        currency = get_admin_text("currency", lang)

        admin_msg = f"👨‍💼 {get_admin_text('admin_panel_title', lang)}\n\n"

        if bot_info:
            admin_msg += f"🤖 {get_admin_text('bot_label', lang)} @{bot_info.get('bot_username', 'N/A')}\n"
           

        admin_msg += f"📊 {get_admin_text('short_stats', lang)}\n"
        admin_msg += f"  👥 {get_admin_text('total_users_label', lang)}: {total_users}\n"
        admin_msg += f"  ✅ {get_admin_text('active_label', lang)}: {active_users}\n"
        admin_msg += f"  ❌ {get_admin_text('removed_label', lang)}: {removed_users}\n"
        admin_msg += f"  💰 {get_admin_text('revenue_label', lang)}: {total_revenue:,.0f} {currency}\n"

        # Admin keyboard with translations
        admin_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"📊 {get_admin_text('btn_stats', lang)}", callback_data="admin_stats")],
            [
                InlineKeyboardButton(text=f"✅ {get_admin_text('btn_active_users', lang)}", callback_data="admin_active_users"),
                InlineKeyboardButton(text=f"❌ {get_admin_text('btn_removed_users', lang)}", callback_data="admin_removed_users")
            ],
            [
                InlineKeyboardButton(text=f"📥 {get_admin_text('btn_users_excel', lang)}", callback_data="admin_users_excel"),
                InlineKeyboardButton(text=f"📥 {get_admin_text('btn_payments_excel', lang)}", callback_data="admin_payments_excel")
            ]
        ])

        await callback.message.edit_text(admin_msg, parse_mode="HTML", reply_markup=admin_keyboard)
        await callback.answer()

    # Echo handler (catch all messages)
    @dp.message()
    async def echo(message):
        logger.info(f"Message from {message.from_user.id}: {message.text[:30] if message.text else 'media'}")
        if message.text:
            await message.answer(f"Siz: {message.text}")
