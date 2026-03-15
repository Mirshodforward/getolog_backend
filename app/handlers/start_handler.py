from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, BufferedInputFile
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from sqlalchemy import text
from app.states import BotCreationStates, BroadcastStates
from app.database import AsyncSessionLocal
from app.crud import get_client_by_user_id, create_client, get_all_client_user_ids, get_users_grouped_by_admin, get_all_active_bots_grouped_by_owner, update_client_language
from aiogram import Bot
from app.config import settings
from app.translations import get_text, get_language_keyboard, _
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import asyncio
import logging

router = Router()
logger = logging.getLogger(__name__)

# Manager Search States
class ManagerStates(StatesGroup):
    waiting_for_client_search = State()
    waiting_for_balance_amount = State()


@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext) -> None:
    """Start command handler"""
    # MUHIM: Har qanday state ni tozalash - bu bot o'chirilgandan keyin /start ishlamasligi muammosini hal qiladi
    await state.clear()
    
    user_id = message.from_user.id
    username = message.from_user.username
    first_name = message.from_user.first_name

    # Check if user is ADMIN/MANAGER
    if user_id in settings.ADMIN_ID:
        # MANAGER ADMIN PANEL
        async with AsyncSessionLocal() as session:
            try:
                # Get global statistics
                total_clients_query = text("SELECT COUNT(*) FROM clients")
                total_clients_result = await session.execute(total_clients_query)
                total_clients = total_clients_result.scalar() or 0

                total_bots_query = text("SELECT COUNT(*) FROM client_bots")
                total_bots_result = await session.execute(total_bots_query)
                total_bots = total_bots_result.scalar() or 0

                active_bots_query = text("SELECT COUNT(*) FROM client_bots WHERE status = 'active'")
                active_bots_result = await session.execute(active_bots_query)
                active_bots = active_bots_result.scalar() or 0

                total_users_query = text("SELECT COUNT(*) FROM users")
                total_users_result = await session.execute(total_users_query)
                total_users = total_users_result.scalar() or 0

                total_revenue_query = text("""
                    SELECT COALESCE(SUM(amount), 0) FROM transactions
                    WHERE role = 'client_topup' AND status = 'approved'
                """)
                total_revenue_result = await session.execute(total_revenue_query)
                total_revenue = float(total_revenue_result.scalar() or 0)

                total_revenue_query_users = text("""
                    SELECT COALESCE(SUM(amount), 0) FROM transactions
                    WHERE role = 'users_topup' AND status = 'approved'
                """)
                total_revenue_result_users = await session.execute(total_revenue_query_users)
                total_revenue_users = float(total_revenue_result_users.scalar() or 0)

            except Exception as e:
                total_clients = 0
                total_bots = 0
                active_bots = 0
                total_users = 0
                total_revenue = 0
                total_revenue_users = 0
        # Manager panel message
        manager_msg = f"<b>MENEJER ADMIN PANEL</b>\n\n"
        manager_msg += f"<b>Umumiy Statistika:</b>\n"
        
        manager_msg += f"  🤖 Jami botlar: {total_bots} ta\n"
        manager_msg += f"  ✅ Faol botlar: {active_bots} ta\n\n"

        manager_msg += f"  👥 Jami clientlar: {total_clients} ta\n"
        manager_msg += f"  💰 Clientlardan daromad: {total_revenue:,.0f} so'm\n\n"

        manager_msg += f"  👤 Jami userlar: {total_users} ta\n"
        manager_msg += f"  💰 Userlardan daromad: {total_revenue_users:,.0f} so'm\n\n"

        # Manager keyboard
        manager_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            
            [
                InlineKeyboardButton(text="📢 Xabar yuborish", callback_data="mgr_broadcast")
            ]
        ])

        await message.answer(manager_msg, parse_mode="HTML", reply_markup=manager_keyboard)
        return

    # REGULAR CLIENT INTERFACE
    async with AsyncSessionLocal() as session:
        # Get or create client
        client = await get_client_by_user_id(session, user_id)

        if not client:
            # New client - show language selection
            await create_client(session, user_id, username=username, language="uz")

            # Language selection keyboard
            lang_buttons = get_language_keyboard()
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=lang_buttons[0]["text"], callback_data=lang_buttons[0]["callback_data"])],
                [InlineKeyboardButton(text=lang_buttons[1]["text"], callback_data=lang_buttons[1]["callback_data"])],
                [InlineKeyboardButton(text=lang_buttons[2]["text"], callback_data=lang_buttons[2]["callback_data"])],
            ])

            await message.answer(
                f"👋 Assalomu alaykum, {first_name}!\n\n"
                f"Getolog botimizga xush kelibsiz! Bu bot orqali siz shaxsiy botlaringizni yaratishingiz mumkin va shaxsiy kanalingizni oson va tez boshqara olasiz.\n\n"
                f"🌐 Tilni tanlang / Выберите язык / Select language:",
                reply_markup=keyboard
            )
            return

        # Existing client - show main menu in their language
        lang = client.language or "uz"

        # Create inline keyboard with translations
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"🤖 {_('btn_create_bot', lang)}", callback_data="create_bot")],
            [InlineKeyboardButton(text=f"📋 {_('btn_my_bots', lang)}", callback_data="my_bots")],
        ])

        await message.answer(f"👋 {_('welcome', lang, name=first_name)}", reply_markup=keyboard)


@router.message(Command("help"))
async def cmd_help(message: Message) -> None:
    """Help command handler"""
    user_id = message.from_user.id

    # Get client language
    async with AsyncSessionLocal() as session:
        client = await get_client_by_user_id(session, user_id)
        lang = client.language if client else "uz"

    help_texts = {
        "uz": "/start - Botni ishga tushirish\n/help - Yordam ko'rish\n/balance - Balansni ko'rish\n/lang - Tilni o'zgartirish",
        "ru": "/start - Запустить бота\n/help - Помощь\n/balance - Проверить баланс\n/lang - Сменить язык",
        "en": "/start - Start the bot\n/help - View help\n/balance - Check balance\n/lang - Change language"
    }

    await message.answer(help_texts.get(lang, help_texts["uz"]))


@router.message(Command("lang"))
async def cmd_lang(message: Message) -> None:
    """Language change command"""
    user_id = message.from_user.id

    # Get client language
    async with AsyncSessionLocal() as session:
        client = await get_client_by_user_id(session, user_id)

    if not client:
        await message.answer("❌ Avval /start bosing.")
        return

    lang = client.language or "uz"

    # Language selection keyboard
    lang_buttons = get_language_keyboard()
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=lang_buttons[0]["text"], callback_data=lang_buttons[0]["callback_data"])],
        [InlineKeyboardButton(text=lang_buttons[1]["text"], callback_data=lang_buttons[1]["callback_data"])],
        [InlineKeyboardButton(text=lang_buttons[2]["text"], callback_data=lang_buttons[2]["callback_data"])],
    ])

    lang_msgs = {
        "uz": "🌐 Tilni tanlang:",
        "ru": "🌐 Выберите язык:",
        "en": "🌐 Select language:"
    }

    await message.answer(lang_msgs.get(lang, lang_msgs["uz"]), reply_markup=keyboard)


@router.message(Command("plan"))
async def cmd_plan(message: Message) -> None:
    """Plan change/view command"""
    from app.config import PLAN_CONFIG
    
    user_id = message.from_user.id

    # Get client
    async with AsyncSessionLocal() as session:
        client = await get_client_by_user_id(session, user_id)

    if not client:
        await message.answer("❌ Avval /start bosing.")
        return

    lang = client.language or "uz"
    current_plan = client.plan_type or "free"

    # Plan selection keyboard
    plans_info = {
        "free": {
            "uz": "🆓 Free (Bepul, 1 bot)",
            "ru": "🆓 Free (Бесплатно, 1 бот)",
            "en": "🆓 Free (Free, 1 bot)"
        },
        "standard": {
            "uz": f"📘 Standard ({PLAN_CONFIG['standard']['price']:,} so'm, 2 bot)",
            "ru": f"📘 Standard ({PLAN_CONFIG['standard']['price']:,} сум, 2 бота)",
            "en": f"📘 Standard ({PLAN_CONFIG['standard']['price']:,} UZS, 2 bots)"
        },
        "biznes": {
            "uz": f"🎯 Biznes ({PLAN_CONFIG['biznes']['price']:,} so'm, 5 bot)",
            "ru": f"🎯 Biznes ({PLAN_CONFIG['biznes']['price']:,} сум, 5 ботов)",
            "en": f"🎯 Biznes ({PLAN_CONFIG['biznes']['price']:,} UZS, 5 bots)"
        }
    }
    
    keyboard_buttons = [
        [InlineKeyboardButton(text=plans_info["free"][lang], callback_data="plan_free")],
        [InlineKeyboardButton(text=plans_info["standard"][lang], callback_data="plan_standard")],
        [InlineKeyboardButton(text=plans_info["biznes"][lang], callback_data="plan_biznes")]
    ]
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_buttons)
    
    msgs = {
        "uz": f"💰 <b>Sizning tarifingiz:</b> <b>{current_plan.upper()}</b>\n\n<b>Tarifni o'zgartirish:</b>",
        "ru": f"💰 <b>Ваш тариф:</b> <b>{current_plan.upper()}</b>\n\n<b>Изменить тариф:</b>",
        "en": f"💰 <b>Your plan:</b> <b>{current_plan.upper()}</b>\n\n<b>Change plan:</b>"
    }
    
    await message.answer(msgs.get(lang, msgs["uz"]), reply_markup=keyboard, parse_mode="HTML")


# ==================== MANAGER ADMIN PANEL HANDLERS ====================

@router.callback_query(F.data == "mgr_clients")
async def show_clients_list(callback: CallbackQuery):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    await callback.answer("📊 Clientlar ro'yxati tayyorlanmoqda...")

    async with AsyncSessionLocal() as session:
        try:
            # Get statistics
            total_clients = (await session.execute(text("SELECT COUNT(*) FROM clients"))).scalar() or 0
            free_clients = (await session.execute(text("SELECT COUNT(*) FROM clients WHERE plan_type IS NULL OR plan_type = 'free'"))).scalar() or 0
            standard_clients = (await session.execute(text("SELECT COUNT(*) FROM clients WHERE plan_type = 'standard'"))).scalar() or 0
            biznes_clients = (await session.execute(text("SELECT COUNT(*) FROM clients WHERE plan_type = 'biznes'"))).scalar() or 0

            # Create Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Clients"

            # Headers
            headers = ["#", "User ID", "Username", "Telefon", "Balans", "Plan", "Til", "Yaratilgan"]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Get all clients
            clients_result = await session.execute(text("""
                SELECT user_id, username, phone_number, balance, plan_type, language, created_at
                FROM clients
                ORDER BY created_at DESC
            """))
            clients = clients_result.fetchall()

            for idx, client in enumerate(clients, 1):
                client_dict = dict(client._mapping)
                ws.append([
                    idx,
                    client_dict.get('user_id'),
                    client_dict.get('username') or 'N/A',
                    client_dict.get('phone_number') or 'N/A',
                    float(client_dict.get('balance') or 0),
                    client_dict.get('plan_type') or 'free',
                    client_dict.get('language') or 'uz',
                    client_dict.get('created_at').strftime('%Y-%m-%d %H:%M') if client_dict.get('created_at') else 'N/A'
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 18

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Send file
            file_name = f"clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_input = BufferedInputFile(excel_file.read(), filename=file_name)

            back_kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="mgr_back")]
            ])

            await callback.message.answer_document(
                document=excel_input,
                caption=f"👥 <b>Clientlar ro'yxati</b>\n\n"
                        f"📊 Jami: <b>{total_clients}</b> ta\n"
                        f"🆓 Free: <b>{free_clients}</b> ta\n"
                        f"� Standard: <b>{standard_clients}</b> ta\n"
                        f"🎯 Biznes: <b>{biznes_clients}</b> ta",
                parse_mode="HTML",
                reply_markup=back_kb
            )

        except Exception as e:
            logger.error(f"Clients export error: {e}")
            await callback.message.answer(f"❌ Xatolik: {e}")

@router.callback_query(F.data == "mgr_bots")
async def show_bots_list(callback: CallbackQuery):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    await callback.answer("📊 Botlar ro'yxati tayyorlanmoqda...")

    async with AsyncSessionLocal() as session:
        try:
            # Get statistics
            total_bots = (await session.execute(text("SELECT COUNT(*) FROM client_bots"))).scalar() or 0
            active_bots = (await session.execute(text("SELECT COUNT(*) FROM client_bots WHERE status = 'active'"))).scalar() or 0
            inactive_bots = total_bots - active_bots

            # Create Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Bots"

            # Headers
            headers = ["#", "Bot nomi", "Bot username", "Egasi ID", "Egasi username", "Kanal ID", "Status", "Yaratilgan"]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Get all bots with owner info
            bots_result = await session.execute(text("""
                SELECT cb.bot_name, cb.bot_username, cb.user_id, c.username, cb.channel_id, cb.status, cb.created_at
                FROM client_bots cb
                LEFT JOIN clients c ON cb.user_id = c.user_id
                ORDER BY cb.created_at DESC
            """))
            bots = bots_result.fetchall()

            for idx, bot in enumerate(bots, 1):
                bot_dict = dict(bot._mapping)
                ws.append([
                    idx,
                    bot_dict.get('bot_name') or 'N/A',
                    bot_dict.get('bot_username') or 'N/A',
                    bot_dict.get('user_id'),
                    bot_dict.get('username') or 'N/A',
                    bot_dict.get('channel_id') or 'N/A',
                    bot_dict.get('status') or 'N/A',
                    bot_dict.get('created_at').strftime('%Y-%m-%d %H:%M') if bot_dict.get('created_at') else 'N/A'
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 18

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Send file
            file_name = f"bots_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_input = BufferedInputFile(excel_file.read(), filename=file_name)

            back_kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="mgr_back")]
            ])

            await callback.message.answer_document(
                document=excel_input,
                caption=f"🤖 <b>Botlar ro'yxati</b>\n\n"
                        f"📊 Jami: <b>{total_bots}</b> ta\n"
                        f"✅ Faol: <b>{active_bots}</b> ta\n"
                        f"⏸ Nofaol: <b>{inactive_bots}</b> ta",
                parse_mode="HTML",
                reply_markup=back_kb
            )

        except Exception as e:
            logger.error(f"Bots export error: {e}")
            await callback.message.answer(f"❌ Xatolik: {e}")

@router.callback_query(F.data == "mgr_users")
async def show_users_list(callback: CallbackQuery):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    await callback.answer("📊 Userlar ro'yxati tayyorlanmoqda...")

    async with AsyncSessionLocal() as session:
        try:
            # Get statistics
            total_users = (await session.execute(text("SELECT COUNT(*) FROM users"))).scalar() or 0
            monthly_users = (await session.execute(text("SELECT COUNT(*) FROM users WHERE duration = '1 oy'"))).scalar() or 0
            yearly_users = (await session.execute(text("SELECT COUNT(*) FROM users WHERE duration = '1 yil'"))).scalar() or 0
            unlimited_users = (await session.execute(text("SELECT COUNT(*) FROM users WHERE duration = 'cheksiz'"))).scalar() or 0
            active_users = (await session.execute(text("SELECT COUNT(*) FROM users WHERE status = 'active'"))).scalar() or 0
            removed_users = (await session.execute(text("SELECT COUNT(*) FROM users WHERE status = 'removed'"))).scalar() or 0

            # Create Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"

            # Headers
            headers = ["#", "User ID", "Username", "Ism", "Muddat", "Status", "Balans", "Bot egasi ID", "Yaratilgan"]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Get all users
            users_result = await session.execute(text("""
                SELECT user_id, username, name, duration, status, balance, admin_id, created_at
                FROM users
                ORDER BY created_at DESC
            """))
            users = users_result.fetchall()

            for idx, user in enumerate(users, 1):
                user_dict = dict(user._mapping)
                ws.append([
                    idx,
                    user_dict.get('user_id'),
                    user_dict.get('username') or 'N/A',
                    user_dict.get('name') or 'N/A',
                    user_dict.get('duration') or 'N/A',
                    user_dict.get('status') or 'N/A',
                    float(user_dict.get('balance') or 0),
                    user_dict.get('admin_id'),
                    user_dict.get('created_at').strftime('%Y-%m-%d %H:%M') if user_dict.get('created_at') else 'N/A'
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 18

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Send file
            file_name = f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_input = BufferedInputFile(excel_file.read(), filename=file_name)

            back_kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="mgr_back")]
            ])

            await callback.message.answer_document(
                document=excel_input,
                caption=f"👤 <b>Userlar ro'yxati</b>\n\n"
                        f"📊 Jami: <b>{total_users}</b> ta\n"
                        f"✅ Faol: <b>{active_users}</b> | ❌ Chiqarilgan: <b>{removed_users}</b>\n\n"
                        f"📋 <b>Tarif bo'yicha:</b>\n"
                        f"  📆 Oylik: <b>{monthly_users}</b> ta\n"
                        f"  📅 Yillik: <b>{yearly_users}</b> ta\n"
                        f"  ♾ Cheksiz: <b>{unlimited_users}</b> ta",
                parse_mode="HTML",
                reply_markup=back_kb
            )

        except Exception as e:
            logger.error(f"Users export error: {e}")
            await callback.message.answer(f"❌ Xatolik: {e}")

@router.callback_query(F.data == "mgr_transactions")
async def show_transactions_list(callback: CallbackQuery):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    await callback.answer("📊 Tranzaksiyalar tayyorlanmoqda...")

    async with AsyncSessionLocal() as session:
        try:
            # Get statistics
            total_trans = (await session.execute(text("SELECT COUNT(*) FROM transactions"))).scalar() or 0
            approved_trans = (await session.execute(text("SELECT COUNT(*) FROM transactions WHERE status = 'approved'"))).scalar() or 0
            pending_trans = (await session.execute(text("SELECT COUNT(*) FROM transactions WHERE status = 'pending'"))).scalar() or 0
            rejected_trans = (await session.execute(text("SELECT COUNT(*) FROM transactions WHERE status = 'rejected'"))).scalar() or 0
            
            # Total amounts
            approved_amount = (await session.execute(text("SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE status = 'approved'"))).scalar() or 0
            pending_amount = (await session.execute(text("SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE status = 'pending'"))).scalar() or 0

            # Get last 10 client_topup transactions
            topup_query = text("""
                SELECT t.id, t.user_id, c.username, t.amount, t.status, t.created_at
                FROM transactions t
                LEFT JOIN clients c ON t.user_id = c.user_id
                WHERE t.role = 'client_topup'
                ORDER BY t.created_at DESC
                LIMIT 10
            """)
            topup_result = await session.execute(topup_query)
            topups = topup_result.fetchall()

            # Build message with last 10 client topups
            msg = f"💳 <b>Oxirgi 10 ta Client To'lov So'rovlari</b>\n\n"

            if not topups:
                msg += "<i>Hozircha to'lov so'rovlari yo'q.</i>\n\n"
            else:
                for idx, trans in enumerate(topups, 1):
                    trans_dict = dict(trans._mapping)
                    status_emoji = "✅" if trans_dict.get('status') == 'approved' else ("❌" if trans_dict.get('status') == 'rejected' else "⏳")
                    msg += f"{idx}. {status_emoji} <b>{float(trans_dict.get('amount', 0)):,.0f} so'm</b>\n"
                    msg += f"   👤 @{trans_dict.get('username') or 'N/A'} (ID: {trans_dict.get('user_id')})\n"
                    if trans_dict.get('created_at'):
                        msg += f"   📅 {trans_dict['created_at'].strftime('%d.%m.%Y %H:%M')}\n"
                    msg += "\n"

            # Create Excel with ALL transactions
            wb = Workbook()
            ws = wb.active
            ws.title = "Transactions"

            # Headers
            headers = ["#", "ID", "Admin ID", "User ID", "Username", "Summa", "Turi", "Status", "Yaratilgan"]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Get all transactions
            all_trans_result = await session.execute(text("""
                SELECT id, admin_id, user_id, username, amount, role, status, created_at
                FROM transactions
                ORDER BY created_at DESC
            """))
            all_transactions = all_trans_result.fetchall()

            for idx, trans in enumerate(all_transactions, 1):
                trans_dict = dict(trans._mapping)
                
                # Translate role
                role_raw = trans_dict.get('role')
                if role_raw == 'client_topup':
                    role_text = "Client to'ldirish"
                elif role_raw == 'plan_purchase':
                    role_text = "Tarif sotib olish"
                elif role_raw == 'users_topup':
                    role_text = "User to'ldirish"
                else:
                    role_text = role_raw or 'N/A'

                # Translate status
                status_raw = trans_dict.get('status')
                if status_raw == 'approved':
                    status_text = "Tasdiqlangan"
                elif status_raw == 'rejected':
                    status_text = "Bekor qilingan"
                elif status_raw == 'pending':
                    status_text = "Kutilmoqda"
                else:
                    status_text = status_raw or 'N/A'

                ws.append([
                    idx,
                    trans_dict.get('id'),
                    trans_dict.get('admin_id') or 'N/A',
                    trans_dict.get('user_id'),
                    trans_dict.get('username') or 'N/A',
                    float(trans_dict.get('amount') or 0),
                    role_text,
                    status_text,
                    trans_dict.get('created_at').strftime('%Y-%m-%d %H:%M') if trans_dict.get('created_at') else 'N/A'
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 18

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Send file
            file_name = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_input = BufferedInputFile(excel_file.read(), filename=file_name)

            back_kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="mgr_back")]
            ])

            await callback.message.answer_document(
                document=excel_input,
                caption=f"{msg}"
                        f"📊 <b>Umumiy Statistika:</b>\n"
                        f"  📋 Jami: <b>{total_trans}</b> ta\n"
                        f"  ✅ Tasdiqlangan: <b>{approved_trans}</b> ta ({float(approved_amount):,.0f} so'm)\n"
                        f"  ⏳ Kutilmoqda: <b>{pending_trans}</b> ta ({float(pending_amount):,.0f} so'm)\n"
                        f"  ❌ Bekor qilingan: <b>{rejected_trans}</b> ta\n\n"
                        f"📥 To'liq jadval Excel faylda",
                parse_mode="HTML",
                reply_markup=back_kb
            )

        except Exception as e:
            logger.error(f"Transactions export error: {e}")
            await callback.message.answer(f"❌ Xatolik: {e}")

@router.callback_query(F.data == "mgr_search")
async def start_client_search(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    await callback.message.edit_text(
        "🔍 <b>Client qidirish</b>\n\n"
        "Client ID yoki username kiriting:\n"
        "(Masalan: 123456789 yoki @username)",
        parse_mode="HTML"
    )
    await state.set_state(ManagerStates.waiting_for_client_search)
    await callback.answer()

@router.message(ManagerStates.waiting_for_client_search)
async def process_client_search(message: Message, state: FSMContext):
    if message.from_user.id not in settings.ADMIN_ID:
        return

    search_term = message.text.strip()

    async with AsyncSessionLocal() as session:
        try:
            # Try to search by user_id or username
            if search_term.isdigit():
                # Search by ID
                client_query = text("SELECT * FROM clients WHERE user_id = :user_id")
                result = await session.execute(client_query, {"user_id": int(search_term)})
            else:
                # Search by username (remove @ if present)
                username = search_term.lstrip('@')
                client_query = text("SELECT * FROM clients WHERE username = :username")
                result = await session.execute(client_query, {"username": username})

            client = result.fetchone()

            if not client:
                await message.answer("❌ Client topilmadi.")
                await state.clear()
                return

            client_dict = dict(client._mapping)
            client_id = client_dict['user_id']

            # Get client's bots
            bots_query = text("SELECT COUNT(*) FROM client_bots WHERE user_id = :user_id")
            bots_result = await session.execute(bots_query, {"user_id": client_id})
            bots_count = bots_result.scalar() or 0

            # Get client's users (from their bots)
            users_query = text("SELECT COUNT(*) FROM users WHERE admin_id = :admin_id")
            users_result = await session.execute(users_query, {"admin_id": client_id})
            users_count = users_result.scalar() or 0

            # Client details
            msg = f"👤 <b>Client Ma'lumotlari</b>\n\n"
            msg += f"🆔 ID: <code>{client_dict.get('user_id')}</code>\n"
            msg += f"👤 Username: @{client_dict.get('username') or 'N/A'}\n"
            msg += f"📞 Telefon: {client_dict.get('phone_number') or 'N/A'}\n"
            msg += f"💰 Balans: {float(client_dict.get('balance') or 0):,.0f} so'm\n"
            msg += f"💎 Plan: {client_dict.get('plan_type') or 'free'}\n\n"
            msg += f"📊 <b>Statistika:</b>\n"
            msg += f"  🤖 Botlar: {bots_count}\n"
            msg += f"  👥 Userlar: {users_count}\n"

            # Action buttons
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="💰➕ Balans qo'shish", callback_data=f"add_balance_{client_id}"),
                    InlineKeyboardButton(text="💰➖ Balans ayirish", callback_data=f"sub_balance_{client_id}")
                ],
                [InlineKeyboardButton(text="🤖 Botlarini ko'rish", callback_data=f"show_client_bots_{client_id}")],
                [InlineKeyboardButton(text="👥 Userlarini ko'rish", callback_data=f"show_client_users_{client_id}")],
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="mgr_back")]
            ])

            await message.answer(msg, parse_mode="HTML", reply_markup=keyboard)
            await state.clear()

        except Exception as e:
            await message.answer(f"❌ Xatolik: {e}")
            await state.clear()

@router.callback_query(F.data.startswith("add_balance_"))
async def add_balance_start(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    client_id = int(callback.data.split("_")[2])
    await state.update_data(client_id=client_id, action="add")

    await callback.message.edit_text(
        "💰 <b>Balans qo'shish</b>\n\n"
        "Qancha summa qo'shmoqchisiz? (so'mda)\n"
        "Masalan: 50000",
        parse_mode="HTML"
    )
    await state.set_state(ManagerStates.waiting_for_balance_amount)
    await callback.answer()

@router.callback_query(F.data.startswith("sub_balance_"))
async def sub_balance_start(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    client_id = int(callback.data.split("_")[2])
    await state.update_data(client_id=client_id, action="subtract")

    await callback.message.edit_text(
        "💰 <b>Balans ayirish</b>\n\n"
        "Qancha summa ayirmoqchisiz? (so'mda)\n"
        "Masalan: 20000",
        parse_mode="HTML"
    )
    await state.set_state(ManagerStates.waiting_for_balance_amount)
    await callback.answer()

@router.message(ManagerStates.waiting_for_balance_amount)
async def process_balance_change(message: Message, state: FSMContext):
    if message.from_user.id not in settings.ADMIN_ID:
        return

    try:
        amount = float(message.text.strip().replace(",", "").replace(" ", ""))
        data = await state.get_data()
        client_id = data['client_id']
        action = data['action']

        async with AsyncSessionLocal() as session:
            if action == "add":
                update_query = text("UPDATE clients SET balance = COALESCE(balance, 0) + :amount WHERE user_id = :user_id")
            else:
                update_query = text("UPDATE clients SET balance = COALESCE(balance, 0) - :amount WHERE user_id = :user_id")

            await session.execute(update_query, {"amount": amount, "user_id": client_id})
            await session.commit()

            # Get new balance
            balance_query = text("SELECT balance FROM clients WHERE user_id = :user_id")
            balance_result = await session.execute(balance_query, {"user_id": client_id})
            new_balance = float(balance_result.scalar() or 0)

        action_text = "qo'shildi" if action == "add" else "ayirildi"
        await message.answer(
            f"✅ <b>Muvaffaqiyatli!</b>\n\n"
            f"Client ID: {client_id}\n"
            f"Summa {action_text}: {amount:,.0f} so'm\n"
            f"Yangi balans: {new_balance:,.0f} so'm",
            parse_mode="HTML"
        )
        await state.clear()

    except ValueError:
        await message.answer("❌ Noto'g'ri summa! Raqam kiriting.")
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
        await state.clear()

@router.callback_query(F.data.startswith("show_client_bots_"))
async def show_client_bots(callback: CallbackQuery):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    client_id = int(callback.data.split("_")[3])

    async with AsyncSessionLocal() as session:
        try:
            bots_query = text("""
                SELECT bot_name, status, channel_id, created_at
                FROM client_bots
                WHERE user_id = :user_id
                ORDER BY created_at DESC
            """)
            result = await session.execute(bots_query, {"user_id": client_id})
            bots = result.fetchall()

            msg = f"🤖 <b>Client botlari (ID: {client_id})</b>\n\n"

            if not bots:
                msg += "Bu clientda botlar yo'q."
            else:
                for idx, bot in enumerate(bots, 1):
                    bot_dict = dict(bot._mapping)
                    status_emoji = "✅" if bot_dict.get('status') == 'active' else "⏸"
                    msg += f"{idx}. {status_emoji} <b>{bot_dict.get('bot_name')}</b>\n"
                    msg += f"   Kanal ID: {bot_dict.get('channel_id', 'N/A')}\n\n"

            back_kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="mgr_back")]
            ])

            await callback.message.edit_text(msg, parse_mode="HTML", reply_markup=back_kb)
            await callback.answer()

        except Exception as e:
            await callback.answer("❌ Xatolik yuz berdi", show_alert=True)

@router.callback_query(F.data.startswith("show_client_users_"))
async def show_client_users(callback: CallbackQuery):
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    client_id = int(callback.data.split("_")[3])

    async with AsyncSessionLocal() as session:
        try:
            users_query = text("""
                SELECT user_id, username, name, duration, status, created_at
                FROM users
                WHERE admin_id = :admin_id
                ORDER BY created_at DESC
                LIMIT 20
            """)
            result = await session.execute(users_query, {"admin_id": client_id})
            users = result.fetchall()

            msg = f"👥 <b>Client userlari (ID: {client_id})</b>\n\n"

            if not users:
                msg += "Bu clientda userlar yo'q."
            else:
                for idx, user in enumerate(users, 1):
                    user_dict = dict(user._mapping)
                    status_emoji = "✅" if user_dict.get('status') == 'active' else "❌"
                    msg += f"{idx}. {status_emoji} <b>{user_dict.get('name', 'N/A')}</b>\n"
                    msg += f"   @{user_dict.get('username') or 'N/A'}\n"
                    msg += f"   Muddat: {user_dict.get('duration', 'N/A')}\n\n"

                if len(users) == 20:
                    msg += "\n<i>Faqat 20ta ko'rsatildi</i>"

            back_kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="mgr_back")]
            ])

            await callback.message.edit_text(msg, parse_mode="HTML", reply_markup=back_kb)
            await callback.answer()

        except Exception as e:
            await callback.answer("❌ Xatolik yuz berdi", show_alert=True)



# ==================== BROADCAST HANDLERS ====================

@router.callback_query(F.data == "mgr_broadcast")
async def start_broadcast(callback: CallbackQuery, state: FSMContext):
    """Broadcast xabar yuborishni boshlash"""
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    # Statistikani olish
    async with AsyncSessionLocal() as session:
        clients_count = (await session.execute(text("SELECT COUNT(*) FROM clients"))).scalar() or 0
        users_count = (await session.execute(text("SELECT COUNT(*) FROM users"))).scalar() or 0
        total_count = clients_count + users_count

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"👥 Faqat Clientlar ({clients_count})", callback_data="broadcast_clients")],
        [InlineKeyboardButton(text=f"👤 Faqat Userlar ({users_count})", callback_data="broadcast_users")],
        [InlineKeyboardButton(text=f"🌐 Hammaga ({total_count})", callback_data="broadcast_all")]
        
    ])

    await callback.message.edit_text(
        "📢 <b>Broadcast Xabar Yuborish</b>\n\n"
        "Kimga xabar yubormoqchisiz?\n\n"
        f"👥 Clientlar: {clients_count} ta\n"
        f"👤 Userlar: {users_count} ta\n"
        f"🌐 Jami: {total_count} ta",
        parse_mode="HTML",
        reply_markup=keyboard
    )
    await callback.answer()


@router.callback_query(F.data.in_({"broadcast_clients", "broadcast_users", "broadcast_all"}))
async def select_broadcast_audience(callback: CallbackQuery, state: FSMContext):
    """Auditoriya tanlash"""
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    audience = callback.data.split("_")[1]  # clients, users, all
    await state.update_data(audience=audience)

    audience_text = {
        "clients": "👥 Clientlar",
        "users": "👤 Userlar",
        "all": "🌐 Hamma"
    }

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Bekor qilish", callback_data="broadcast_cancel")]
    ])

    await callback.message.edit_text(
        f"📢 <b>Broadcast: {audience_text.get(audience, audience)}</b>\n\n"
        "Endi xabar yuboring:\n\n"
        "• Oddiy matn yuboring\n"
        "• Yoki rasm + caption yuboring\n\n"
        "<i>Xabar barcha tanlangan foydalanuvchilarga yuboriladi</i>",
        parse_mode="HTML",
        reply_markup=keyboard
    )
    await state.set_state(BroadcastStates.entering_message)
    await callback.answer()


@router.callback_query(F.data == "broadcast_cancel")
async def cancel_broadcast(callback: CallbackQuery, state: FSMContext):
    """Broadcastni bekor qilish"""
    if callback.from_user.id not in settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    await state.clear()

    # Rasm yoki matn xabarni o'chirish va yangi xabar yuborish
    try:
        await callback.message.delete()
    except Exception:
        pass

    await callback.message.answer(
        "❌ Broadcast bekor qilindi.\n\n"
        "/start - Bosh menyuga qaytish"
    )
    await callback.answer()


@router.message(BroadcastStates.entering_message, F.photo)
async def receive_broadcast_photo(message: Message, state: FSMContext):
    """Rasm + caption qabul qilish"""
    if message.from_user.id != settings.ADMIN_ID:
        return

    photo = message.photo[-1]  # Eng katta o'lchamdagi rasm
    caption = message.caption or ""

    await state.update_data(
        message_type="photo",
        photo_id=photo.file_id,
        caption=caption
    )

    data = await state.get_data()
    audience = data.get("audience", "all")

    audience_text = {
        "clients": "👥 Clientlar",
        "users": "👤 Userlar",
        "all": "🌐 Hamma"
    }

    # Statistikani olish
    async with AsyncSessionLocal() as session:
        if audience == "clients":
            count = (await session.execute(text("SELECT COUNT(*) FROM clients"))).scalar() or 0
        elif audience == "users":
            count = (await session.execute(text("SELECT COUNT(*) FROM users"))).scalar() or 0
        else:
            clients_count = (await session.execute(text("SELECT COUNT(*) FROM clients"))).scalar() or 0
            users_count = (await session.execute(text("SELECT COUNT(*) FROM users"))).scalar() or 0
            count = clients_count + users_count

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Tasdiqlash va yuborish", callback_data="broadcast_confirm")],
        [InlineKeyboardButton(text="❌ Bekor qilish", callback_data="broadcast_cancel")]
    ])

    await message.answer_photo(
        photo=photo.file_id,
        caption=f"📢 <b>Broadcast tasdiqlash</b>\n\n"
                f"📌 Auditoriya: <b>{audience_text.get(audience, audience)}</b>\n"
                f"👥 Qabul qiluvchilar: <b>{count} ta</b>\n\n"
                f"<b>Xabar:</b>\n{caption or '<i>Caption yo`q</i>'}",
        parse_mode="HTML",
        reply_markup=keyboard
    )
    await state.set_state(BroadcastStates.confirming_broadcast)


@router.message(BroadcastStates.entering_message, F.text)
async def receive_broadcast_text(message: Message, state: FSMContext):
    """Matn xabarni qabul qilish"""
    if message.from_user.id != settings.ADMIN_ID:
        return

    await state.update_data(
        message_type="text",
        text=message.text
    )

    data = await state.get_data()
    audience = data.get("audience", "all")

    audience_text = {
        "clients": "👥 Clientlar",
        "users": "👤 Userlar",
        "all": "🌐 Hamma"
    }

    # Statistikani olish
    async with AsyncSessionLocal() as session:
        if audience == "clients":
            count = (await session.execute(text("SELECT COUNT(*) FROM clients"))).scalar() or 0
        elif audience == "users":
            count = (await session.execute(text("SELECT COUNT(*) FROM users"))).scalar() or 0
        else:
            clients_count = (await session.execute(text("SELECT COUNT(*) FROM clients"))).scalar() or 0
            users_count = (await session.execute(text("SELECT COUNT(*) FROM users"))).scalar() or 0
            count = clients_count + users_count

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Tasdiqlash va yuborish", callback_data="broadcast_confirm")],
        [InlineKeyboardButton(text="❌ Bekor qilish", callback_data="broadcast_cancel")]
    ])

    await message.answer(
        f"📢 <b>Broadcast tasdiqlash</b>\n\n"
        f"📌 Auditoriya: <b>{audience_text.get(audience, audience)}</b>\n"
        f"👥 Qabul qiluvchilar: <b>{count} ta</b>\n\n"
        f"<b>Xabar:</b>\n{message.text}",
        parse_mode="HTML",
        reply_markup=keyboard
    )
    await state.set_state(BroadcastStates.confirming_broadcast)


@router.callback_query(BroadcastStates.confirming_broadcast, F.data == "broadcast_confirm")
async def confirm_and_send_broadcast(callback: CallbackQuery, state: FSMContext):
    """Broadcastni tasdiqlash va yuborish"""
    if callback.from_user.id != settings.ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q", show_alert=True)
        return

    data = await state.get_data()
    audience = data.get("audience", "all")
    message_type = data.get("message_type", "text")

    # Eski xabarni o'chirish va yangi status xabari yuborish
    try:
        await callback.message.delete()
    except Exception:
        pass

    status_msg = await callback.message.answer(
        "⏳ <b>Xabarlar yuborilmoqda...</b>\n\n"
        "Iltimos kuting...",
        parse_mode="HTML"
    )
    await callback.answer()

    main_bot = callback.bot
    success_count = 0
    fail_count = 0
    blocked_count = 0

    # Helper funksiya - xabar yuborish
    async def send_broadcast_message(bot_instance, user_id):
        if message_type == "photo":
            await bot_instance.send_photo(
                chat_id=user_id,
                photo=data.get("photo_id"),
                caption=data.get("caption"),
                parse_mode="HTML"
            )
        else:
            await bot_instance.send_message(
                chat_id=user_id,
                text=data.get("text"),
                parse_mode="HTML"
            )

    async with AsyncSessionLocal() as session:
        # 1. CLIENTLARGA - asosiy bot orqali
        if audience in ["clients", "all"]:
            client_ids = await get_all_client_user_ids(session)
            for user_id in client_ids:
                try:
                    await send_broadcast_message(main_bot, user_id)
                    success_count += 1
                    await asyncio.sleep(0.05)
                except Exception as e:
                    error_msg = str(e).lower()
                    if "blocked" in error_msg or "deactivated" in error_msg or "not found" in error_msg:
                        blocked_count += 1
                    else:
                        fail_count += 1
                    logger.warning(f"Broadcast error for client {user_id}: {e}")

        # 2. USERLARGA - ularni kirgan boti orqali
        if audience in ["users", "all"]:
            # Userlarni admin_id bo'yicha guruhlash
            users_by_admin = await get_users_grouped_by_admin(session)
            # Har bir admin uchun faol bot tokenini olish
            bots_by_owner = await get_all_active_bots_grouped_by_owner(session)

            # Har bir admin va uning userlari uchun
            for admin_id, user_ids in users_by_admin.items():
                bot_token = bots_by_owner.get(admin_id)

                if not bot_token:
                    # Bu admin uchun faol bot yo'q - skip
                    fail_count += len(user_ids)
                    logger.warning(f"No active bot for admin {admin_id}, skipping {len(user_ids)} users")
                    continue

                # Vaqtinchalik bot instance yaratish
                temp_bot = None
                try:
                    temp_bot = Bot(token=bot_token)

                    for user_id in user_ids:
                        try:
                            await send_broadcast_message(temp_bot, user_id)
                            success_count += 1
                            await asyncio.sleep(0.05)
                        except Exception as e:
                            error_msg = str(e).lower()
                            if "blocked" in error_msg or "deactivated" in error_msg or "not found" in error_msg:
                                blocked_count += 1
                            else:
                                fail_count += 1
                            logger.warning(f"Broadcast error for user {user_id} via bot {admin_id}: {e}")

                except Exception as e:
                    fail_count += len(user_ids)
                    logger.error(f"Error creating bot instance for admin {admin_id}: {e}")
                finally:
                    if temp_bot:
                        await temp_bot.session.close()

    # Natija
    total = success_count + fail_count + blocked_count
    await status_msg.edit_text(
        f"✅ <b>Broadcast yakunlandi!</b>\n\n"
        f"📊 <b>Statistika:</b>\n"
        f"  ✅ Muvaffaqiyatli: {success_count}\n"
        f"  🚫 Bloklangan: {blocked_count}\n"
        f"  ❌ Xatolik: {fail_count}\n"
        f"  📝 Jami: {total}\n\n"
        f"/start - Bosh menyuga qaytish",
        parse_mode="HTML"
    )

    await state.clear()
    logger.info(f"Broadcast completed: {success_count}/{total} successful")


# Echo handler olib tashlandi - bu keraksiz va muammolar keltirib chiqaradi
# Barcha xabarlar tegishli state handler lari orqali qayta ishlanadi
